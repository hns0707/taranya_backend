"""
Import client invoice Excel into UAT customers + sale_invoices for CRM testing.

Usage (from Backend/ecom_backend):
  pip install openpyxl
  python manage.py import_client_invoices --file ../../imports/Invoices.xlsx --dry-run
  python manage.py import_client_invoices --file ../../imports/Invoices.xlsx

Excel sheet \"Invoices\" columns:
  Invoice Number, Invoice Date, Bill To Name, Phone, Address,
  Total Amount, Paid Amount, Pending Amount, Status, Deleted,
  Birthday, Anniversary, Refferal, Family
"""
from __future__ import annotations

import re
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path

from django.contrib.auth.hashers import make_password
from django.core.management.base import BaseCommand, CommandError
from django.db import transaction
from django.utils import timezone
from django.utils.crypto import get_random_string

from shared.models import Customer, SaleInvoice


def _norm_phone(raw) -> str:
    digits = re.sub(r"\D", "", str(raw or ""))
    if len(digits) >= 10:
        return digits[-10:]
    return digits


def _as_date(val) -> date | None:
    if val is None or val == "":
        return None
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val
    text = str(val).strip()
    for fmt in ("%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y", "%Y/%m/%d"):
        try:
            return datetime.strptime(text[:10], fmt).date()
        except ValueError:
            continue
    return None


def _as_decimal(val) -> Decimal:
    if val is None or val == "":
        return Decimal("0")
    try:
        return Decimal(str(val).replace(",", "").strip())
    except (InvalidOperation, ValueError):
        return Decimal("0")


def _as_bool_deleted(val) -> bool:
    if val is None:
        return False
    if isinstance(val, bool):
        return val
    s = str(val).strip().lower()
    return s in {"1", "true", "yes", "y"}


def _norm_status(raw: str) -> str:
    s = (raw or "PENDING").strip().upper()
    if s in {SaleInvoice.STATUS_PAID, SaleInvoice.STATUS_PARTIAL, SaleInvoice.STATUS_PENDING}:
        return s
    if "PAID" in s:
        return SaleInvoice.STATUS_PAID
    if "PARTIAL" in s:
        return SaleInvoice.STATUS_PARTIAL
    return SaleInvoice.STATUS_PENDING


def _row_map(header: tuple) -> dict[str, int]:
    mapping = {}
    aliases = {
        "invoice_number": ("invoice number", "invoice_no", "invoice no"),
        "invoice_date": ("invoice date", "date"),
        "bill_to_name": ("bill to name", "customer name", "name"),
        "phone": ("phone", "mobile", "bill to phone"),
        "address": ("address", "bill to address"),
        "total_amount": ("total amount", "total"),
        "paid_amount": ("paid amount", "paid"),
        "pending_amount": ("pending amount", "pending", "due"),
        "status": ("status",),
        "deleted": ("deleted", "is_deleted"),
        "birthday": ("birthday", "date of birth", "dob"),
        "anniversary": ("anniversary", "anniversary date"),
        "referral": ("refferal", "referral", "referred by"),
        "family": ("family", "family group"),
    }
    lower = {str(h or "").strip().lower(): i for i, h in enumerate(header)}
    for key, names in aliases.items():
        for name in names:
            if name in lower:
                mapping[key] = lower[name]
                break
    required = ["invoice_number", "bill_to_name", "phone"]
    missing = [k for k in required if k not in mapping]
    if missing:
        raise CommandError(f"Missing required columns: {missing}. Found: {list(header)}")
    return mapping


class Command(BaseCommand):
    help = "Import Invoices.xlsx into customers + sale_invoices (UAT / CRM testing)."

    def add_arguments(self, parser):
        parser.add_argument(
            "--file",
            type=str,
            default="",
            help="Path to Invoices.xlsx (default: <repo>/imports/Invoices.xlsx)",
        )
        parser.add_argument(
            "--sheet",
            type=str,
            default="Invoices",
            help="Worksheet name (default: Invoices)",
        )
        parser.add_argument(
            "--dry-run",
            action="store_true",
            help="Parse and report without writing to the database.",
        )
        parser.add_argument(
            "--update-customers",
            action="store_true",
            help="Update existing customers' DOB/anniversary/family when blank.",
        )

    def handle(self, *args, **options):
        try:
            import openpyxl
        except ImportError as exc:
            raise CommandError("openpyxl is required. Run: pip install openpyxl") from exc

        file_opt = (options.get("file") or "").strip()
        if file_opt:
            path = Path(file_opt).expanduser().resolve()
        else:
            # .../shared/management/commands/this.py → repo root = parents[5]
            repo_root = Path(__file__).resolve().parents[5]
            path = repo_root / "imports" / "Invoices.xlsx"
            if not path.exists():
                path = Path(__file__).resolve().parents[3] / "imports" / "Invoices.xlsx"

        if not path.exists():
            raise CommandError(f"File not found: {path}")

        dry_run = bool(options["dry_run"])
        update_customers = bool(options["update_customers"])
        sheet_name = options["sheet"]

        self.stdout.write(f"Reading {path} sheet={sheet_name!r} dry_run={dry_run}")

        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        if sheet_name not in wb.sheetnames:
            wb.close()
            raise CommandError(f"Sheet {sheet_name!r} not found. Available: {wb.sheetnames}")

        ws = wb[sheet_name]
        rows_iter = ws.iter_rows(values_only=True)
        try:
            header = next(rows_iter)
        except StopIteration as exc:
            wb.close()
            raise CommandError("Empty worksheet") from exc

        col = _row_map(header)

        stats = {
            "rows": 0,
            "customers_created": 0,
            "customers_reused": 0,
            "customers_updated": 0,
            "invoices_created": 0,
            "invoices_skipped": 0,
            "errors": 0,
        }
        error_samples: list[str] = []
        customer_cache: dict[str, Customer] = {}

        def cell(row, key, default=None):
            idx = col.get(key)
            if idx is None:
                return default
            return row[idx] if idx < len(row) else default

        @transaction.atomic
        def process_all():
            for row in rows_iter:
                if not row or all(v is None or str(v).strip() == "" for v in row):
                    continue
                stats["rows"] += 1
                try:
                    inv_no = str(cell(row, "invoice_number") or "").strip()
                    name = str(cell(row, "bill_to_name") or "").strip()
                    phone = _norm_phone(cell(row, "phone"))
                    if not inv_no or not name or len(phone) < 10:
                        raise ValueError(f"incomplete row invoice={inv_no!r} name={name!r} phone={phone!r}")

                    address = str(cell(row, "address") or "").strip() or "—"
                    inv_date = _as_date(cell(row, "invoice_date")) or timezone.localdate()
                    total = _as_decimal(cell(row, "total_amount"))
                    paid = _as_decimal(cell(row, "paid_amount"))
                    pending = _as_decimal(cell(row, "pending_amount"))
                    status = _norm_status(str(cell(row, "status") or ""))
                    is_deleted = _as_bool_deleted(cell(row, "deleted"))
                    dob = _as_date(cell(row, "birthday"))
                    anniversary = _as_date(cell(row, "anniversary"))
                    family = str(cell(row, "family") or "").strip()[:150] or None
                    referral_name = str(cell(row, "referral") or "").strip()

                    if phone in customer_cache:
                        customer = customer_cache[phone]
                        stats["customers_reused"] += 1
                    else:
                        existing = Customer.objects.filter(mobile=phone).first()
                        if not existing:
                            # also match 91-prefixed stored mobiles
                            existing = Customer.objects.filter(mobile__endswith=phone).first()

                        if existing:
                            customer = existing
                            stats["customers_reused"] += 1
                            if update_customers and not dry_run:
                                changed = []
                                if dob and not customer.date_of_birth:
                                    customer.date_of_birth = dob
                                    changed.append("date_of_birth")
                                if anniversary and not customer.anniversary_date:
                                    customer.anniversary_date = anniversary
                                    changed.append("anniversary_date")
                                if family and not (customer.family_group or "").strip():
                                    customer.family_group = family
                                    changed.append("family_group")
                                if changed:
                                    customer.save(update_fields=changed + ["system_updated_at"])
                                    stats["customers_updated"] += 1
                        else:
                            if dry_run:
                                customer = Customer(
                                    full_name=name,
                                    mobile=phone,
                                    date_of_birth=dob,
                                    anniversary_date=anniversary,
                                    family_group=family,
                                )
                            else:
                                password = get_random_string(10)
                                customer = Customer.objects.create(
                                    full_name=name,
                                    mobile=phone,
                                    email=None,
                                    date_of_birth=dob,
                                    anniversary_date=anniversary,
                                    family_group=family,
                                    password_hash=make_password(password),
                                    is_active=True,
                                )
                                # simple unique-ish code for POS/CRM
                                code = f"IMP-{customer.id:05d}"
                                if not Customer.objects.filter(customer_code=code).exists():
                                    customer.customer_code = code
                                    customer.save(update_fields=["customer_code"])
                            stats["customers_created"] += 1

                        # optional referred_by by exact full_name (skip vague "Ashish" if many matches)
                        if referral_name and not dry_run and customer.pk and not customer.referred_by_id:
                            matches = list(
                                Customer.objects.filter(full_name__iexact=referral_name).exclude(pk=customer.pk)[:2]
                            )
                            if len(matches) == 1:
                                customer.referred_by = matches[0]
                                customer.save(update_fields=["referred_by", "system_updated_at"])

                        customer_cache[phone] = customer

                    if SaleInvoice.objects.filter(invoice_number=inv_no).exists():
                        stats["invoices_skipped"] += 1
                        continue

                    if dry_run:
                        stats["invoices_created"] += 1
                        continue

                    SaleInvoice.objects.create(
                        invoice_number=inv_no,
                        customer=customer if getattr(customer, "pk", None) else None,
                        bill_to_name=name[:150],
                        bill_to_phone=phone[:15],
                        bill_to_address=address,
                        invoice_date=inv_date,
                        total_amount=total,
                        paid_amount=paid,
                        pending_amount=pending,
                        status=status,
                        is_deleted=is_deleted,
                    )
                    stats["invoices_created"] += 1
                except Exception as exc:  # noqa: BLE001 — collect per-row errors
                    stats["errors"] += 1
                    if len(error_samples) < 15:
                        error_samples.append(str(exc))

            if dry_run:
                transaction.set_rollback(True)

        process_all()
        wb.close()

        self.stdout.write(self.style.SUCCESS("Import finished"))
        for k, v in stats.items():
            self.stdout.write(f"  {k}: {v}")
        if error_samples:
            self.stdout.write(self.style.WARNING("Sample errors:"))
            for e in error_samples:
                self.stdout.write(f"  - {e}")

        if dry_run:
            self.stdout.write(self.style.WARNING("Dry-run only — no data was written."))
        else:
            self.stdout.write(
                self.style.SUCCESS(
                    "Customers + invoices are ready for CRM testing "
                    "(dashboard conversion uses sale_invoices linked by customer)."
                )
            )
